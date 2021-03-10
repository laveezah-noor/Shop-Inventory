from tkinter import *
from tkinter import ttk
from openpyxl import *
from tkinter import messagebox



class Products(Frame):
    def __init__(self,*args,**kwargs):
        Frame.__init__(self,*args,**kwargs)
        self.label = Label(self, text="Products",
                        font=("times new roman", 20, "bold"))
        self.label.grid(row=0, column=1)
        
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(1, weight=1)
        
        self.productprice = StringVar()
        self.productname = StringVar()
        self.productqty = StringVar()
       
        Manage_Frame = Frame(self, bd=4, relief=RIDGE, bg="crimson")
        Manage_Frame.place(x=20, y=100, width=450, height=600)

        m_title = Label(Manage_Frame, text="Manage Products", bg="crimson", fg="white",
                        font=("times new roman", 20, "bold"))
        m_title.grid(row=0, columnspan=2, pady=20)

        lbl_productname= Label(Manage_Frame, text="Product Name", bg="crimson", fg="white", font=("times new roman", 18, "bold"))
        lbl_productname.grid(row=1, column=0, padx=10, pady=10, sticky="w")
        txt_productname= Entry(Manage_Frame, textvariable=self.productname, font=("times new roman", 18, "bold"), bd=5,
                         relief=GROOVE)
        txt_productname.grid(row=1, column=1, padx=10, pady=10, sticky="w")

        lbl_productprice= Label(Manage_Frame, text="Product Price", bg="crimson", fg="white", font=("times new roman", 18, "bold"))
        lbl_productprice.grid(row=2, column=0, padx=10, pady=10, sticky="w")
        txt_productprice= Entry(Manage_Frame, textvariable=self.productprice, font=("times new roman", 18, "bold"), bd=5,
                         relief=GROOVE)
        txt_productprice.grid(row=2, column=1, padx=10, pady=10, sticky="w")

        lbl_productqty= Label(Manage_Frame, text="Quantity", bg="crimson", fg="white", font=("times new roman", 18, "bold"))
        lbl_productqty.grid(row=3, column=0, padx=10, pady=10, sticky="w")
        txt_productqty= Entry(Manage_Frame, textvariable=self.productqty, font=("times new roman", 18, "bold"), bd=5,
                         relief=GROOVE)
        txt_productqty.grid(row=3, column=1, padx=10, pady=10, sticky="w")
        
        lbl_add_button = Button(Manage_Frame, text="Add", bg="crimson", fg="white",
                            font=("times new roman", 18, "bold"), command=self.add_data)
        lbl_add_button.grid(row=8, column=0, padx=10, pady=10, sticky="w")
        
        lbl_clear_button = Button(Manage_Frame, text="Clear", bg="crimson", fg="white",
                            font=("times new roman", 18, "bold"), command=self.clear)
        lbl_clear_button.grid(row=8, column=1, padx=10, pady=10, sticky="w")
         
        Detail_Frame = Frame(self, bd=4, relief=RIDGE, bg="crimson")
        Detail_Frame.place(x=500, y=100, width=750, height=580)
 
        btn_show = Button(Detail_Frame, text="Show Data", font=("times new roman", 13), bd=5,
                            relief=GROOVE, command=self.show_data, bg="crimson")
        btn_show.grid(row=0, column=2, padx=5, pady=5, sticky="w")
        
        # Table Frame
        Table_Frame = Frame(Detail_Frame, bd=4, relief=RIDGE, bg="crimson")
        Table_Frame.place(x=20, y=60)
 
        scroll_x = Scrollbar(Table_Frame, orient=HORIZONTAL)
        scroll_y = Scrollbar(Table_Frame, orient=VERTICAL)
 
        self.Product_table = ttk.Treeview(Table_Frame,
           columns=("productid", "productname", "productprice", "productqty"),
           xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)
        scroll_x.pack(side=BOTTOM, fill=X)
        scroll_y.pack(side=RIGHT, fill=Y)
        scroll_x.config(command=self.Product_table.xview)
        scroll_y.config(command=self.Product_table.yview)
        self.Product_table.heading("productid", text="Product#")
        self.Product_table.heading("productname", text="Name")
        self.Product_table.heading("productprice", text="Price")
        self.Product_table.heading("productqty", text="Quantity")
        self.Product_table['show'] = 'headings'  # removing extra index col at begining
 
        # setting up widths of cols
        self.Product_table.column("productid", width=100)
        self.Product_table.column("productname", width=100)
        self.Product_table.column("productprice", width=100)
        self.Product_table.column("productqty", width=100)
        self.Product_table.pack(fill=BOTH, expand=1)  # fill both is used to fill cols around the frame
        self.Product_table.bind("<ButtonRelease-1>", self.get_cursor)  # this is an event to select row
 
    def clear(self):
        self.amount.set("")
        self.name_var.set("")
        self.email_var.set("")
        self.product.set("")
        self.qty.set("")
        self.dob_var.set("")

    def get_cursor(self, evnt):
        cursor_row = self.Customer_table.focus()
        content = self.Customer_table.item(cursor_row)
        row = content['values']
        self.amount.set(row[0])
        self.name_var.set(row[1])
        self.email_var.set(row[2])
        self.product.set(row[3])
        self.qty.set(row[4])
        self.dob_var.set(row[5])

    def show_data(self):
        workbook = load_workbook('data.xlsx')
        sheet = workbook.active
        row_max = sheet.max_row
        col_max = sheet.max_column
        for i in self.Customer_table.get_children():
            self.Customer_table.delete(i)
        for i in range(1,row_max):
            no = sheet.cell(column=1, row=i+1).value
            name = sheet.cell(column=2, row=i+1).value
            email = sheet.cell(column=3, row=i+1).value
            contact = sheet.cell(column=4, row=i+1).value
            orders = sheet.cell(column=5, row=i+1).value
            qty = sheet.cell(column=6, row=i+1).value
            print(no, name, email, contact, orders, qty)
            self.Customer_table.insert('', 'end', text=no, values=(no, name, email, contact, orders, qty))
        messagebox.showinfo("successfull", "Record has been updated.")

    def delete_data(self):
        self.fetch_data()
        self.clear()
        messagebox.showinfo("successfull", "Record has been deleted.")

    def search_data(self):

        sql = "SELECT * FROM Orders WHERE roll_no = %s"
        adr = self.search_txt.get()

        val = cur.execute(sql, adr)
        if (not val):
            messagebox.showinfo("No", "Not availabe!")

        # cur.execute("SELECT * FROM Orders WHERE"+str(params)+"LIKE %s ", ("%"+str(params2)+"%",))
        # cur.execute("SELECT * from Orders WHERE"+str(self.search_by.get())+"LIKE '%"+str(self.search_txt.get())+"%'")
        rows = cur.fetchall()
        if (len(rows) != 0):
            self.Customer_table.delete(*self.Customer_table.get_children())
            for row in rows:
                self.Customer_table.insert('', END, values=row)

            con.commit()
        con.close()
    
    def add_data(self):
        workbook = load_workbook(filename='data.xlsx')
        sheet = workbook.active
        print(sheet)
        roll = self.amount.get()
        name = self.name_var.get()
        email = self.email_var.get()
        gender = self.product.get()
        contact = self.qty.get()
        dob = self.dob_var.get()
        print(roll, name, email, gender, contact, dob)
        data = [roll, name, email, gender, contact, dob]
        sheet.append(data)
        workbook.save(filename='data.xlsx')

        


        # self.productId = productId
        # self.productPrice = productPrice
        # self.productName = productName

    # def addProduct(self, productId, productPrice, productName):
    #     self.userSheet.append([productId, productPrice, productName])

    # def showProducts(self):
    #     row_max = self.userSheet.max_row
    #     col_max = self.userSheet.max_column
