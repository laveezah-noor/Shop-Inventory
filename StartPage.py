from tkinter import *
from Login import Login
from openpyxl import *


class StartPage(Frame):

    def __init__(self, parent, controller):
        Frame.__init__(self, parent)

        workbook = load_workbook('data.xlsx')
        sheet = workbook.active
        for i in range(1,sheet.max_row):
            sheet.cell(row=i+1, column=5, value=0)
        workbook.save(filename='data.xlsx')

        label1 = Label(self, text="Let's Get Started!")
        label2 = Label(self, text="With Your Own Inventory")
        label1.pack(padx=10, pady=10)
        label2.pack(padx=10, pady=10)
        login = Button(self, text="Get Started", command=lambda:controller.show_frame(Login))
        login.pack()
