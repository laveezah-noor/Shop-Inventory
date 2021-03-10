from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from openpyxl import *


class Customers(Frame):
    def __init__(self,*args,**kwargs):
        Frame.__init__(self,*args,**kwargs)
        self.label = Label(self, text="Hi This is Tab2")
        self.label.grid(row=1,column=0,padx=10,pady=10)  
        
        self.contact_var = StringVar()
        self.name_var = StringVar()
        self.email_var = StringVar()
       
        Manage_Frame = Frame(self, bd=4, relief=RIDGE, bg="pink")
        Manage_Frame.place(x=10, y=50, width=450, height=600)

        m_title = Label(Manage_Frame, text="Manage Customers", bg="pink", fg="white",
                        font=("times new roman", 20, "bold"))
        m_title.grid(row=0, columnspan=2, pady=10)

        lbl_name = Label(Manage_Frame, text="Name", bg="crimson", fg="white", font=("times new roman", 18, "bold"))
        lbl_name.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        txt_name = Entry(Manage_Frame, textvariable=self.name_var, font=("times new roman", 18, "bold"), bd=5,
                         relief=GROOVE)
        txt_name.grid(row=1, column=1, padx=10, pady=5, sticky="w")

        lbl_email = Label(Manage_Frame, text="Email", bg="crimson", fg="white", font=("times new roman", 18, "bold"))
        lbl_email.grid(row=2, column=0, padx=10, pady=5, sticky="w")
        txt_email = Entry(Manage_Frame, textvariable=self.email_var, font=("times new roman", 18, "bold"), bd=5,
                         relief=GROOVE)
        txt_email.grid(row=2, column=1, padx=10, pady=5, sticky="w")

        lbl_contact = Label(Manage_Frame, text="Contact", bg="crimson", fg="white", font=("times new roman", 18, "bold"))
        lbl_contact.grid(row=3, column=0, padx=10, pady=5, sticky="w")
        txt_contact = Entry(Manage_Frame, textvariable=self.contact_var, font=("times new roman", 18, "bold"), bd=5,
                          relief=GROOVE)
        txt_contact.grid(row=3, column=1, padx=10, pady=5, sticky="w")

        lbl_add_button = Button(Manage_Frame, text="Add", bg="pink", fg="white",
                            font=("times new roman", 18, "bold"), command=self.add_data)
        lbl_add_button.grid(row=8, column=0, padx=10, pady=5, sticky="w")
        
        lbl_clear_button = Button(Manage_Frame, text="Clear", bg="pink", fg="white",
                            font=("times new roman", 18, "bold"), command=self.clear)
        lbl_clear_button.grid(row=8, column=1, padx=10, pady=5, sticky="w")
         
        Detail_Frame = Frame(self, bd=4, relief=RIDGE, bg="pink")
        Detail_Frame.place(x=250, y=50, width=750, height=580)
 
        btn_show = Button(Detail_Frame, text="Show Data", font=("times new roman", 13), bd=5,
                            relief=GROOVE, command=self.show_data, bg="pink")
        btn_show.grid(row=0, column=2, padx=5, pady=5, sticky="w")
        
        # Table Frame
        Table_Frame = Frame(Detail_Frame, bd=4, relief=RIDGE, bg="pink")
        Table_Frame.place(x=20, y=60)
 
        scroll_x = Scrollbar(Table_Frame, orient=HORIZONTAL)
        scroll_y = Scrollbar(Table_Frame, orient=VERTICAL)
 
        self.Customer_table = ttk.Treeview(Table_Frame,
           columns=("no", "name", "email", "contact", "orders"),
           xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)
        scroll_x.pack(side=BOTTOM, fill=X)
        scroll_y.pack(side=RIGHT, fill=Y)
        scroll_x.config(command=self.Customer_table.xview)
        scroll_y.config(command=self.Customer_table.yview)
        self.Customer_table.heading("no", text="ID#")
        self.Customer_table.heading("name", text="Name")
        self.Customer_table.heading("email", text="Email Address")
        self.Customer_table.heading("contact", text="Contact")
        self.Customer_table.heading("orders", text="Orders")
        self.Customer_table['show'] = 'headings'  # removing extra index col at begining
 
        # setting up widths of cols
        self.Customer_table.column("no", width=100)
        self.Customer_table.column("name", width=100)
        self.Customer_table.column("email", width=100)
        self.Customer_table.column("contact", width=100)
        self.Customer_table.column("orders", width=100)
        self.Customer_table.pack(fill=BOTH, expand=1)  # fill both is used to fill cols around the frame
        self.Customer_table.bind("<ButtonRelease-1>", self.get_cursor)  # this is an event to select row
 
    def clear(self):
        self.amount.set("")
        self.name_var.set("")
        self.email_var.set("")
        self.product.set("")
        self.qty.set("")
        self.dob_var.set("")

    def get_cursor(self, evnt):
        cursor_row = self.Customer_table.focus()
        content = self.Customer_table.item(cursor_row)
        row = content['values']
        self.amount.set(row[0])
        self.name_var.set(row[1])
        self.email_var.set(row[2])
        self.product.set(row[3])
        self.qty.set(row[4])
        self.dob_var.set(row[5])

    def show_data(self):
        workbook = load_workbook('data.xlsx')
        sheet = workbook.active
        row_max = sheet.max_row
        col_max = sheet.max_column
        for i in self.Customer_table.get_children():
            self.Customer_table.delete(i)
        for i in range(1,row_max):
            no = sheet.cell(column=1, row=i+1).value
            name = sheet.cell(column=2, row=i+1).value
            email = sheet.cell(column=3, row=i+1).value
            contact = sheet.cell(column=4, row=i+1).value
            orders = sheet.cell(column=5, row=i+1).value
            qty = sheet.cell(column=6, row=i+1).value
            print(no, name, email, contact, orders, qty)
            self.Customer_table.insert('', 'end', text=no, values=(no, name, email, contact, orders, qty))
        messagebox.showinfo("successfull", "Record has been updated.")

    def delete_data(self):
        self.fetch_data()
        self.clear()
        messagebox.showinfo("successfull", "Record has been deleted.")

    def search_data(self):

        sql = "SELECT * FROM Orders WHERE roll_no = %s"
        adr = self.search_txt.get()

        val = cur.execute(sql, adr)
        if (not val):
            messagebox.showinfo("No", "Not availabe!")

        # cur.execute("SELECT * FROM Orders WHERE"+str(params)+"LIKE %s ", ("%"+str(params2)+"%",))
        # cur.execute("SELECT * from Orders WHERE"+str(self.search_by.get())+"LIKE '%"+str(self.search_txt.get())+"%'")
        rows = cur.fetchall()
        if (len(rows) != 0):
            self.Customer_table.delete(*self.Customer_table.get_children())
            for row in rows:
                self.Customer_table.insert('', END, values=row)

            con.commit()
        con.close()
    
    def add_data(self):
        workbook = load_workbook(filename='data.xlsx')
        sheet = workbook.active
        print(sheet)
        roll = self.amount.get()
        name = self.name_var.get()
        email = self.email_var.get()
        gender = self.product.get()
        contact = self.qty.get()
        dob = self.dob_var.get()
        print(roll, name, email, gender, contact, dob)
        data = [roll, name, email, gender, contact, dob]
        sheet.append(data)
        workbook.save(filename='data.xlsx')
