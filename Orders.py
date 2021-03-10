from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from openpyxl import *


class Orders(Frame):
    def __init__(self,*args,**kwargs):
        Frame.__init__(self,*args,**kwargs)

        self.amount = StringVar()
        self.name_var = StringVar()
        self.email_var = StringVar()
        self.product = StringVar()
        self.qty = StringVar()
        self.dob_var = StringVar()
       
        Manage_Frame = Frame(self, bd=4, relief=RIDGE, bg="crimson")
        Manage_Frame.place(x=20, y=100, width=450, height=600)

        m_title = Label(Manage_Frame, text="Manage Orders", bg="crimson", fg="white",
                        font=("times new roman", 20, "bold"))
        m_title.grid(row=0, columnspan=2, pady=20)

        lbl_amount = Label(Manage_Frame, text="Amount", bg="crimson", fg="white", font=("times new roman", 18, "bold"))
        lbl_amount.grid(row=1, column=0, padx=20, pady=10, sticky="w")
        txt_amount = Entry(Manage_Frame, textvariable=self.amount, font=("times new roman", 18, "bold"), bd=5,
                         relief=GROOVE)
        txt_amount.grid(row=1, column=1, padx=20, pady=10, sticky="w")

        lbl_product = Label(Manage_Frame, text="Product", bg="crimson", fg="white", font=("times new roman", 18, "bold"))
        lbl_product.grid(row=2, column=0, padx=20, pady=10, sticky="w")
        combo_product = ttk.Combobox(Manage_Frame, textvariable=self.product, font=("times new roman", 13, "bold"),
                                    state="readonly")
        combo_product['values'] = ("Male", "Female", "Other")
        combo_product.grid(row=2, column=1, padx=20, pady=10, sticky="w")

        lbl_customer = Label(Manage_Frame, text="Customer", bg="crimson", fg="white", font=("times new roman", 18, "bold"))
        lbl_customer.grid(row=3, column=0, padx=20, pady=10, sticky="w")
        combo_customer = ttk.Combobox(Manage_Frame, textvariable=self.product, font=("times new roman", 13, "bold"),
                                    state="readonly")
        combo_customer['values'] = ("Male", "Female", "Other")
        combo_customer.grid(row=3, column=1, padx=20, pady=10, sticky="w")

        lbl_qty = Label(Manage_Frame, text="Quantity", bg="crimson", fg="white",
                            font=("times new roman", 18, "bold"))
        lbl_qty.grid(row=5, column=0, padx=20, pady=10, sticky="w")
        txt_qty = Entry(Manage_Frame, textvariable=self.qty, font=("times new roman", 18, "bold"), bd=5,
                            relief=GROOVE)
        txt_qty.grid(row=5, column=1, padx=20, pady=10, sticky="w")

        lbl_add_button = Button(Manage_Frame, text="Add", bg="crimson", fg="white",
                            font=("times new roman", 18, "bold"), command=self.add_data)
        lbl_add_button.grid(row=8, column=0, padx=20, pady=10, sticky="w")
        
        lbl_clear_button = Button(Manage_Frame, text="Clear", bg="crimson", fg="white",
                            font=("times new roman", 18, "bold"), command=self.clear)
        lbl_clear_button.grid(row=8, column=1, padx=20, pady=10, sticky="w")
         
        Detail_Frame = Frame(self, bd=4, relief=RIDGE, bg="crimson")
        Detail_Frame.place(x=500, y=100, width=750, height=580)
 
        btn_show = Button(Detail_Frame, text="Show Data", font=("times new roman", 13), bd=5,
                            relief=GROOVE, command=self.show_data, bg="crimson")
        btn_show.grid(row=0, column=2, padx=5, pady=5, sticky="w")
        
        # Table Frame
        Table_Frame = Frame(Detail_Frame, bd=4, relief=RIDGE, bg="crimson")
        Table_Frame.place(x=20, y=60)
 
        scroll_x = Scrollbar(Table_Frame, orient=HORIZONTAL)
        scroll_y = Scrollbar(Table_Frame, orient=VERTICAL)
 
        self.Order_table = ttk.Treeview(Table_Frame,
           columns=("order", "date", "customer", "product", "amount", "qty"),
           xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)
        scroll_x.pack(side=BOTTOM, fill=X)
        scroll_y.pack(side=RIGHT, fill=Y)
        scroll_x.config(command=self.Order_table.xview)
        scroll_y.config(command=self.Order_table.yview)
        self.Order_table.heading("order", text="Order#")
        self.Order_table.heading("date", text="Date")
        self.Order_table.heading("customer", text="Customer Name")
        self.Order_table.heading("product", text="Product Name")
        self.Order_table.heading("amount", text="Amount")
        self.Order_table.heading("qty", text="Quantity")
        self.Order_table['show'] = 'headings'  # removing extra index col at begining
 
        # setting up widths of cols
        self.Order_table.column("order", width=100)
        self.Order_table.column("date", width=100)
        self.Order_table.column("customer", width=100)
        self.Order_table.column("product", width=100)
        self.Order_table.column("amount", width=100)
        self.Order_table.column("qty", width=100)
        self.Order_table.pack(fill=BOTH, expand=1)  # fill both is used to fill cols around the frame
        self.Order_table.bind("<ButtonRelease-1>", self.get_cursor)  # this is an event to select row
 
    def clear(self):
        self.amount.set("")
        self.name_var.set("")
        self.email_var.set("")
        self.product.set("")
        self.qty.set("")
        self.dob_var.set("")

    def get_cursor(self, evnt):
        cursor_row = self.Order_table.focus()
        content = self.Order_table.item(cursor_row)
        row = content['values']
        self.amount.set(row[0])
        self.name_var.set(row[1])
        self.email_var.set(row[2])
        self.product.set(row[3])
        self.qty.set(row[4])
        self.dob_var.set(row[5])

    def show_data(self):
        workbook = load_workbook('data.xlsx')
        sheet = workbook.active
        row_max = sheet.max_row
        col_max = sheet.max_column
        for i in self.Order_table.get_children():
            self.Order_table.delete(i)
        for i in range(1,row_max):
            order = sheet.cell(column=1, row=i+1).value
            date = sheet.cell(column=2, row=i+1).value
            customer = sheet.cell(column=3, row=i+1).value
            product = sheet.cell(column=4, row=i+1).value
            amount = sheet.cell(column=5, row=i+1).value
            qty = sheet.cell(column=6, row=i+1).value
            print(order, date, customer, product, amount, qty)
            self.Order_table.insert('', 'end', text=order, values=(order, date, customer, product, amount, qty))
        messagebox.showinfo("successfull", "Record has been updated.")

    def delete_data(self):
        self.fetch_data()
        self.clear()
        messagebox.showinfo("successfull", "Record has been deleted.")

    def search_data(self):

        sql = "SELECT * FROM Orders WHERE roll_no = %s"
        adr = self.search_txt.get()

        val = cur.execute(sql, adr)
        if (not val):
            messagebox.showinfo("No", "Not availabe!")

        # cur.execute("SELECT * FROM Orders WHERE"+str(params)+"LIKE %s ", ("%"+str(params2)+"%",))
        # cur.execute("SELECT * from Orders WHERE"+str(self.search_by.get())+"LIKE '%"+str(self.search_txt.get())+"%'")
        rows = cur.fetchall()
        if (len(rows) != 0):
            self.Order_table.delete(*self.Order_table.get_children())
            for row in rows:
                self.Order_table.insert('', END, values=row)

            con.commit()
        con.close()
    
    def add_data(self):
        workbook = load_workbook(filename='data.xlsx')
        sheet = workbook.active
        print(sheet)
        roll = self.amount.get()
        name = self.name_var.get()
        email = self.email_var.get()
        gender = self.product.get()
        contact = self.qty.get()
        dob = self.dob_var.get()
        print(roll, name, email, gender, contact, dob)
        data = [roll, name, email, gender, contact, dob]
        sheet.append(data)
        workbook.save(filename='data.xlsx')
