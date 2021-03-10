from tkinter import *
from tkinter import ttk
from Home import Home
from Register import Register
import json
from openpyxl import load_workbook


class Login(Frame):

    def __init__(self, parent, controller):
        Frame.__init__(self, parent)

        temp_email = StringVar()
        temp_password = StringVar()

        # Labels
        Label(self, text="Please enter your details below to Login", font=('Calibri', 12)).grid(row=0, sticky=N, pady=10, padx=2)
        Label(self, text="Email", font=('Calibri', 12)).grid(row=1, column=0, sticky=W, padx=2)
        Label(self, text="Password", font=('Calibri', 12)).grid(row=2, column=0, sticky=W, padx=2)

        # Entries
        Entry(self, textvariable=temp_email).grid(row=1, column=0)
        Entry(self, textvariable=temp_password, show="*").grid(row=2, column=0)

        Button(self, text="Login", command=lambda: login(temp_email,temp_password)).grid(row=3, column=0)
        notif = Label(self, font=('Calibri',12))
        notif.grid(row=6,sticky=N,pady=10)

        def login(temp_email, temp_password):
            print(temp_email.get(), temp_password.get())
            workbook = load_workbook(filename="data.xlsx")
            userSheet = workbook['Users']
            row_max = userSheet.max_row 
            for i in range(row_max):
                name = userSheet.cell(row=(i+1), column=1).value
                email = userSheet.cell(row=(i+1), column=2).value
                password = userSheet.cell(row=(i+1), column=3).value
                file = userSheet.cell(row=(i+1), column=4).value
                status = userSheet.cell(row=(i+1), column=5).value
                print("Login==> ", name, email, password, file, status)
                if temp_email.get() == email:
                    if temp_password.get() == password:
                        userSheet.cell(row=(i+1), column=5).value = TRUE
                        print(userSheet.cell(row=(i+1), column=5).value)
                        workbook.save(filename='data.xlsx')
                        user = load_workbook(filename=file)
                        print("Login==>",user.active, user.sheetnames)
                        user.save(filename=file)
                        controller.show_frame(Home)
                        break
                    else:
                        notif.config(fg="red", text="Incorrect Email or Password", font=('Ariel', 8))
                else:
                    notif.config(fg="red", text="No User", font=('Ariel', 8))

        def register():
            Register(self)

        Label(self, text="Not registered?", font=('Calibri', 8)).grid(row=4, column=0, sticky=(S, W), padx=0, pady=5)
        Button(self, text="Register", command=register).grid(row=4, column=0)


