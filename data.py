# from openpyxl import load_workbook
from openpyxl import Workbook
# from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
# wb = Workbook()
# workbook = load_workbook(filename="demo.xlsx")
workbook = Workbook()

sheet = workbook.active
sheet.title = 'Users'
# sheet1 = workbook['Orders']
# import string, datetime
# import random, math
# def id_generator(size=6, chars=string.ascii_uppercase + string.digits):
#     return ''.join(random.choice(chars) for _ in range(size))
# def add_Order(orderid, productid, productName, customerid, customername,  amount,  qty):
#     orderdate = datetime.datetime.now().strftime("%x")
#     data = [orderid, orderdate, productid, productName, customerid, customername,  amount,  qty]
#     sheet.append(data)
# add_Order(id_generator(), )
#
# def add_Product(productid, productprice, producttype):
#     data = [orderid, orderdate, productid, productName, customerid, customername,  amount,  qty]
#     sheet.append(data)
#
# def add_Customer(name, email, telephone, address):
#     data = [name, email, telephone, address]
#     sheet.append(data)
sheet["A1"] = "User Name"
sheet["B1"] = "User Email"
sheet["C1"] = "User Password"
sheet["D1"] = "User File"
sheet["E1"] = "User Status"

# font = Font(name='Calibri',
#     size=12,
#     bold=True,
#     italic=False,
#     vertAlign=None,
#     underline='none',
#     strike=False,
#     color='FF000000')
# sheet["A1"].font = font
# sheet["B1"].font = font
# sheet["C1"].font = font
# sheet.title =
# products_sheet = workbook["Products"]
# row_max = sheet.max_row
# col_max = sheet.max_column

# for i in range(row_max):
#     print(sheet.cell(row=(i+1), column=1).value)

# data = list(data)
# print(data)
# sheet.append(data)
# workbook.create_sheet('User1')
# sheet.delete_rows(1)
# for sheet in workbook.sheetnames:
#     std=workbook.get_sheet_by_name(sheet)
#     workbook.remove_sheet(std)
#     print(sheet)
# print(workbook.sheetnames)
workbook.save(filename="data.xlsx")
# print(workbook.sheetnames)
# print(workbook.active)

# print(workbook.active.title)
