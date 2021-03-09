from tkinter import *
from tkinter import ttk
from openpyxl import *


class Home(Frame):

    def __init__(self, parent, controller):
        Frame.__init__(self, parent)

        workbook = load_workbook('data.xlsx')
        print("Home==> ",workbook.active)
        sheet = workbook.active
        row_max = sheet.max_row 
        col_max = sheet.max_column  
        for i in range(1,row_max):
            for j in range(1,col_max):
                text = sheet.cell(row=(i), column=(j)).value
                print("Home==> ",text)
                label = Label(self, text=text)
                label.pack(padx=20, pady=20)
        
            
        label = Label(self, text="Home")
        label.pack(padx=10, pady=10)
        page_one = Button(self, text="Products", command=lambda:controller.show_frame(Product))
        page_one.pack()
        page_two = Button(self, text="Customers", command=lambda:controller.show_frame(Customer))
        page_two.pack()
        page_three = Button(self, text="Orders", command=lambda:controller.show_frame(Order))
        page_three.pack()
        # mainframe = ttk.Frame(master, padding="3 3 12 12")
        # mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
        # master.columnconfigure(0, weight=1)
        # master.rowconfigure(0, weight=1)
        #
        # ttk.Button(mainframe, text="Get Started", command=None).grid(column=3, row=3, sticky=(N, W, E, S))
        # ttk.Label(mainframe, text="Shop Inventory").grid(column=3, row=1, sticky=W)
        # #
        # for child in mainframe.winfo_children():
        #     child.grid_configure(padx=15, pady=15)

        # feet_entry.focus()

        # master.bind("<Return>", self.calculate)


# root = Tk()
# Home(root)
# root.mainloop()
