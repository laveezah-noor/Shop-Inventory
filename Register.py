#imports
from tkinter import *
from openpyxl import *


class Register:
    def __init__(self, parent):
        temp_name = StringVar()
        temp_email = StringVar()
        # temp_gender = StringVar()
        temp_password = StringVar()

        # Register Screen
        register_screen = Toplevel(parent)
        register_screen.title('Register')

        #Labels
        Label(register_screen, text="Please enter your details below to register", font=('Calibri',12)).grid(row=0,sticky=N,pady=10)
        Label(register_screen, text="Name", font=('Calibri',12)).grid(row=1,sticky=W)
        Label(register_screen, text="Email", font=('Calibri',12)).grid(row=2,sticky=W)
        # Label(register_screen, text="Gender", font=('Calibri',12)).grid(row=3,sticky=W)
        Label(register_screen, text="Password", font=('Calibri',12)).grid(row=4,sticky=W)
        notif = Label(register_screen, font=('Calibri',12))
        notif.grid(row=6,sticky=N,pady=10)

        #Entries
        Entry(register_screen,textvariable=temp_name).grid(row=1,column=0)
        Entry(register_screen,textvariable=temp_email).grid(row=2,column=0)
        # Entry(register_screen,textvariable=temp_gender).grid(row=3,column=0)
        Entry(register_screen,textvariable=temp_password,show="*").grid(row=4,column=0)

        def finish_reg():
            name = temp_name.get()
            email = temp_email.get()
            password = temp_password.get()
            path = "data.xlsx"
            workbook = load_workbook(filename=path)
            userSheet = workbook['Users']

            if name == "" or email == "" or password == "":
                notif.config(fg="red", text="All fields required * ", font=('Ariel', 8))
            else:
                notif.config(text="", font=('Ariel', 8))

            i = 1
            col_max = userSheet.max_column
            while col_max >= i:
                userName = userSheet.cell(row=(i+1), column=1).value
                userEmail = userSheet.cell(row=(i+1), column=2).value
                print(userName, userEmail)
                if name == userName or email == userEmail:
                    notif.config(fg="red", font=('Calibri', 8), text="Account already exists")
                    break
                else:
                    i += 1
            else:
                newFile = name + '.xlsx'
                newFile = newFile.replace(" ", "").lower()
                print(newFile)
                data = [name, email, password, newFile, FALSE]
                row_max = userSheet.max_row
                print(row_max)
                for col_ind, title in enumerate(data, start=1):
                    userSheet.cell(row=row_max+1, column=col_ind,value=title)
                workbook.save(filename="data.xlsx")
                newWorkbook = Workbook()
                newWs = newWorkbook.active
                newWs.title = 'Profile'
                newWs.append(['UserName', 'UserEmail', 'Password', 'File', 'Status'])
                newWorkbook.create_sheet('Products')
                newWorkbook.create_sheet('Customers')
                newWorkbook.create_sheet('Orders')
                print(data)
                row_max = newWs.max_row
                print(row_max)
                for col_ind, title in enumerate(data, start=1):
                    newWs.cell(row=row_max+1, column=col_ind,value=title)
                newWorkbook.save(filename=(newFile))
                notif.config(fg='green', text="Account has been created")
                register_screen.destroy()

        #Buttons
        Button(register_screen, text="Register", command=finish_reg, font=('Calibri',12)).grid(row=5,sticky=N,pady=10)
